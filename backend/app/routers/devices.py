import asyncio
import io
import logging
import re
from datetime import datetime, timezone, timedelta
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.database import get_db
from app.models.device import Device
from app.schemas.device import (
    DeviceCreate, DeviceUpdate, DeviceResponse, DeviceListResponse,
    DeviceDiscoverRequest, DeviceDiscoverResponse, DiscoveredDevice,
    DeviceBatchCreate, BatchDeviceDeleteRequest,
)
from app.services.discovery_service import discover_device, collect_entity_components, pick_chassis_serial
from app.services.credential_service import protect_device_secrets, reveal_secret
from app.routers.auth import current_user
from app.services.audit_service import record_audit, get_client_ip

router = APIRouter(prefix="/devices", tags=["设备管理"])

# 实体 MIB 组件中排除物理接口/端口实体（GigabitEthernet1/0/1 等）
_PORT_ENTITY_RE = re.compile(
    r"(GigabitEthernet|Ten-GigabitEthernet|FortyGigE|HundredGigE|XGigabitEthernet|"
    r"TwentyGigE|Twenty-FiveGigE|Ethernet\d|M-GigabitEthernet)", re.I
)

# Vendor mapping: English filter -> DB values to match
VENDOR_MAP = {
    "Huawei": ["华为", "Huawei", "huawei"],
    "H3C": ["H3C", "h3c"],
    "Cisco": ["Cisco", "cisco", "思科"],
    "Juniper": ["Juniper", "juniper"],
    "Arista": ["Arista", "arista"],
    "锐捷": ["锐捷", "Ruijie"],
    "深信服": ["深信服", "Sangfor"],
}

# 设备列表可排序字段（前端表头点击排序）
SORTABLE_FIELDS = {
    "name": Device.name,
    "ip": Device.ip,
    "vendor": Device.vendor,
    "model": Device.model,
    "device_type": Device.device_type,
    "status": Device.status,
    "cpu_usage": Device.cpu_usage,
    "memory_usage": Device.memory_usage,
    "last_seen": Device.last_seen,
}


def _device_filter_conditions(vendor: str | None, device_type: str | None, status: str | None, keyword: str | None) -> list:
    """设备列表 / 导出的公共筛选条件。"""
    conds = []
    if vendor:
        # Map English vendor name to possible DB values
        vendor_variants = VENDOR_MAP.get(vendor, [vendor])
        vendor_filters = [Device.vendor.ilike(f"%{v}") for v in vendor_variants] + \
                         [Device.vendor.ilike(f"{v}%") for v in vendor_variants] + \
                         [Device.vendor == v for v in vendor_variants]
        conds.append(or_(*vendor_filters))
    if device_type:
        conds.append(Device.device_type == device_type)
    if status:
        conds.append(Device.status == status)
    if keyword:
        conds.append(or_(Device.name.ilike(f"%{keyword}%"), Device.ip.ilike(f"%{keyword}%")))
    return conds


# 资产清单导出列（中文表头 -> 模型字段）
DEVICE_EXPORT_COLUMNS = [
    ("设备名称", "name"),
    ("IP 地址", "ip"),
    ("厂商", "vendor"),
    ("型号", "model"),
    ("序列号", "serial_number"),
    ("设备类型", "device_type"),
    ("分组", "group_name"),
    ("位置", "location"),
    ("状态", "status"),
    ("CPU 使用率(%)", "cpu_usage"),
    ("内存使用率(%)", "memory_usage"),
    ("最后在线时间", "last_seen"),
    ("创建时间", "created_at"),
    ("保修到期", "warranty_expire"),
    ("停止销售(EOS)", "eos_date"),
    ("停止支持(EOL)", "eol_date"),
]

DEVICE_TYPE_NAMES = {
    "router": "路由器", "switch": "交换机", "firewall": "防火墙",
    "load_balancer": "负载均衡", "server": "服务器",
}
STATUS_NAMES = {"online": "在线", "warning": "告警", "offline": "离线", "unknown": "未知"}


@router.get("", response_model=DeviceListResponse)
async def list_devices(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=300),
    vendor: str | None = None,
    device_type: str | None = None,
    status: str | None = None,
    keyword: str | None = None,
    sort: str | None = Query(None, description="排序字段：name/ip/vendor/model/device_type/status/cpu_usage/memory_usage/last_seen"),
    order: str = Query("asc", pattern="^(asc|desc)$"),
    db: AsyncSession = Depends(get_db),
):
    query = select(Device)
    count_query = select(func.count(Device.id))

    conds = _device_filter_conditions(vendor, device_type, status, keyword)
    query = query.where(*conds)
    count_query = count_query.where(*conds)

    total = (await db.execute(count_query)).scalar()
    offset = (page - 1) * page_size
    sort_col = SORTABLE_FIELDS.get(sort) if sort else None
    if sort_col is not None:
        sort_expr = sort_col.asc() if order == "asc" else sort_col.desc()
        query = query.order_by(sort_expr, Device.id)
    else:
        query = query.order_by(Device.id)
    result = await db.execute(query.offset(offset).limit(page_size))
    items = result.scalars().all()

    return DeviceListResponse(total=total, items=[DeviceResponse.model_validate(d) for d in items])


@router.get("/export")
async def export_devices(
    vendor: str | None = None,
    device_type: str | None = None,
    status: str | None = None,
    keyword: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    """按当前筛选条件导出全部设备为 Excel 资产清单（不受分页限制）。"""
    conds = _device_filter_conditions(vendor, device_type, status, keyword)
    result = await db.execute(select(Device).where(*conds).order_by(Device.id))
    devices = result.scalars().all()

    tz_8 = timezone(timedelta(hours=8))

    def fmt_dt(v):
        if v is None:
            return ""
        if isinstance(v, datetime):
            v = v.astimezone(tz_8)
        return v.strftime("%Y-%m-%d %H:%M:%S")

    wb = Workbook()
    ws = wb.active
    ws.title = "资产清单"
    headers = ["序号"] + [col[0] for col in DEVICE_EXPORT_COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F6FEB")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, d in enumerate(devices, 1):
        row = [idx]
        for _, field in DEVICE_EXPORT_COLUMNS:
            val = getattr(d, field, None)
            if field == "device_type":
                val = DEVICE_TYPE_NAMES.get(val, val or "")
            elif field == "status":
                val = STATUS_NAMES.get(val, val or "")
            elif isinstance(val, datetime):
                val = fmt_dt(val)
            elif isinstance(val, float):
                val = round(val, 1)
            row.append(val if val is not None else "")
        ws.append(row)

    # 自适应列宽（限宽 8~32 字符）
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 8), 32)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"设备资产清单_{datetime.now(tz_8).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.get("/{device_id}/interfaces")
async def get_device_interfaces(
    device_id: int,
    top: int = 10,
    db: AsyncSession = Depends(get_db),
):
    """获取设备流量最大的物理接口列表（实时 SNMP 双采样计算）。

    - 返回接口按最大利用率（上行/下行取大）降序，默认前 10 个；
    - 需要设备 SNMP 读权限（community 在设备信息中）。
    """
    from app.services.interface_traffic_service import collect_interface_traffic

    result = await db.execute(select(Device).where(Device.id == device_id))
    device = result.scalar_one_or_none()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    community = reveal_secret(device.snmp_community) or "aiops"
    try:
        interfaces = await collect_interface_traffic(
            device.ip, community=community,
        )
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"接口流量采集失败: {e}")

    interfaces.sort(key=lambda x: x["max_util"], reverse=True)
    top_n = max(1, min(top, 50))
    return {
        "device_id": device_id,
        "total": len(interfaces),
        "interfaces": interfaces[:top_n],
    }


@router.get("/{device_id}", response_model=DeviceResponse)
async def get_device(device_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Device).where(Device.id == device_id))
    device = result.scalar_one_or_none()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    return DeviceResponse.model_validate(device)


@router.post("", response_model=DeviceResponse, status_code=201)
async def create_device(data: DeviceCreate, actor: dict = Depends(current_user), db: AsyncSession = Depends(get_db)):
    current_count = (await db.execute(select(func.count(Device.id)))).scalar() or 0
    if current_count >= settings.MAX_DEVICES:
        raise HTTPException(
            status_code=400,
            detail=f"设备数量已达上限（{settings.MAX_DEVICES} 台），无法继续添加。如需扩容请联系管理员。",
        )
    # 同一 IP 只能添加一台设备（历史数据可能已有重复 IP，用 first 保证不崩溃）
    if data.ip:
        existing = (await db.execute(
            select(Device).where(Device.ip == data.ip)
        )).scalars().first()
        if existing:
            raise HTTPException(
                status_code=409,
                detail=f"设备（{data.ip}）已存在（ID {existing.id}，名称 {existing.name}），请勿重复添加",
            )
    device = Device(**protect_device_secrets(data.model_dump()))
    db.add(device)
    await db.commit()
    await db.refresh(device)
    # 自动通过 SNMP 补全厂商/型号/序列号/类型等（仅填充用户留空字段）
    await _enrich_device(db, device)
    await record_audit(db, actor, "device", "create", f"添加设备 {device.name}({device.ip})")
    return DeviceResponse.model_validate(device)


@router.put("/{device_id}", response_model=DeviceResponse)
async def update_device(device_id: int, data: DeviceUpdate, actor: dict = Depends(current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Device).where(Device.id == device_id))
    device = result.scalar_one_or_none()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    # 修改 IP 时校验唯一性：不能改成其他设备已占用的 IP
    if data.ip is not None and data.ip != device.ip:
        existing = (await db.execute(
            select(Device).where(Device.ip == data.ip, Device.id != device_id)
        )).scalars().first()
        if existing:
            raise HTTPException(
                status_code=409,
                detail=f"设备（{data.ip}）已存在（ID {existing.id}，名称 {existing.name}），无法使用该 IP",
            )

    update_data = protect_device_secrets(data.model_dump(exclude_unset=True))
    # 凭据类字段：空字符串视为"未修改"（前端编辑表单提交空值时不覆盖真实凭据）
    for key, value in update_data.items():
        if key in ("snmp_community", "mgmt_username", "mgmt_password") and value in (None, ""):
            continue
        setattr(device, key, value)

    await db.commit()
    await db.refresh(device)
    await record_audit(db, actor, "device", "update", f"更新设备 {device.name}({device.ip})")
    return DeviceResponse.model_validate(device)


@router.delete("/{device_id}", status_code=204)
async def delete_device(device_id: int, actor: dict = Depends(current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Device).where(Device.id == device_id))
    device = result.scalar_one_or_none()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    # 先清理 NO ACTION 外键关联数据（与批量删除一致），避免 NotNullViolation
    await _cleanup_device_relations(db, [device.id])
    await db.delete(device)
    await db.commit()
    # 清理告警引擎内存状态，避免残留导致误判
    from app.services.alert_rule_engine import cleanup_device_state
    cleanup_device_state(device_id)
    await record_audit(db, actor, "device", "delete", f"删除设备 {device.name}({device.ip})")


@router.post("/batch-delete", response_model=DeviceListResponse)
async def batch_delete_devices(data: BatchDeviceDeleteRequest, db: AsyncSession = Depends(get_db), user: dict = Depends(current_user)):
    """批量删除设备。

    支持两种模式：
    - device_ids 非空：删除指定 ID 列表中的设备；
    - delete_all=True：删除全部设备（需传空的 device_ids，仅管理员可执行）。

    删除前先清理外键关联数据（alerts / security_events / compliance_checks /
    topology_links 为 NO ACTION 约束，直接删设备会触发 NotNullViolation），
    否则返回 422 并提示。
    """
    if data.delete_all and user["role"] != "admin":
        raise HTTPException(status_code=403, detail="清空全部设备需要管理员权限")
    if data.delete_all:
        result = await db.execute(select(Device).order_by(Device.id))
    else:
        ids = list(set(data.device_ids or []))
        if not ids:
            raise HTTPException(status_code=422, detail="未选择要删除的设备")
        result = await db.execute(select(Device).where(Device.id.in_(ids)))

    targets = result.scalars().all()
    if not targets:
        raise HTTPException(status_code=404, detail="所选设备不存在")

    target_ids = [d.id for d in targets]
    await _cleanup_device_relations(db, target_ids)

    for d in targets:
        await db.delete(d)
    await db.commit()

    # 清理告警引擎内存状态
    from app.services.alert_rule_engine import cleanup_device_state
    for did in target_ids:
        cleanup_device_state(did)

    # 删除后返回剩余设备（delete_all 则已清空）
    remaining = (await db.execute(select(Device).order_by(Device.id))).scalars().all()
    names = ", ".join(d.name for d in targets) or "-"
    await record_audit(db, user, "device", "batch_delete", f"批量删除 {len(target_ids)} 台设备（{names[:100]}）")
    return DeviceListResponse(total=len(remaining), items=[DeviceResponse.model_validate(d) for d in remaining])


async def _cleanup_device_relations(db: AsyncSession, device_ids: list[int]) -> None:
    """删除设备前清理 NO ACTION 外键关联数据（单删/批量删除共用）。

    - Alert.device_id / ComplianceCheck.device_id / TopologyLink(source|target)_device_id
      为 NOT NULL + NO ACTION，直接删设备会触发 NotNullViolation → 先删除关联行；
    - SecurityEvent.device_id 为 nullable，置 NULL 即可。
    """
    if not device_ids:
        return
    try:
        from sqlalchemy import delete as sa_delete
        from app.models.alert import Alert
        from app.models.device_component import DeviceComponent
        from app.models.p3_security import ComplianceCheck, SecurityEvent
        from app.models.topology_link import TopologyLink
        from app.models.config_backup import ConfigBackup, BackupSchedule
        from app.models.p2_baseline import MetricBaseline, PredictionResult, DeviceHealthScore
        from app.models.inspection import InspectionDeviceResult

        await db.execute(sa_delete(Alert).where(Alert.device_id.in_(device_ids)))
        await db.execute(sa_delete(DeviceComponent).where(DeviceComponent.device_id.in_(device_ids)))
        await db.execute(sa_delete(ComplianceCheck).where(ComplianceCheck.device_id.in_(device_ids)))
        await db.execute(sa_delete(TopologyLink).where(
            (TopologyLink.source_device_id.in_(device_ids)) |
            (TopologyLink.target_device_id.in_(device_ids))
        ))
        await db.execute(
            SecurityEvent.__table__.update()
            .where(SecurityEvent.device_id.in_(device_ids))
            .values(device_id=None)
        )
        # 以下表 device_id 均为 NOT NULL：虽然 DB 外键是 CASCADE，
        # 但 SQLAlchemy ORM 删除 device 时会先把关联行 FK 置 NULL → NotNullViolation，
        # 因此必须显式删除关联行（或按 CASCADE 语义由 DB 处理前先清掉）。
        await db.execute(sa_delete(ConfigBackup).where(ConfigBackup.device_id.in_(device_ids)))
        await db.execute(sa_delete(BackupSchedule).where(BackupSchedule.device_id.in_(device_ids)))
        await db.execute(sa_delete(MetricBaseline).where(MetricBaseline.device_id.in_(device_ids)))
        await db.execute(sa_delete(PredictionResult).where(PredictionResult.device_id.in_(device_ids)))
        await db.execute(sa_delete(DeviceHealthScore).where(DeviceHealthScore.device_id.in_(device_ids)))
        await db.execute(sa_delete(InspectionDeviceResult).where(InspectionDeviceResult.device_id.in_(device_ids)))
        await db.flush()
    except Exception:
        # 关联清理失败不致命：交由下方约束校验兜底
        await db.rollback()


@router.post("/sync-all")
async def sync_all_devices(
    user: dict = Depends(current_user),
):
    """一键同步所有设备信息（并发 SNMP 重新发现，覆盖厂商/型号/类型/序列号/名称）。

    每台设备使用独立 DB 会话并发执行（受限并发，避免打爆设备/网络），
    单台失败不中断整批，返回逐台结果。
    """
    from app.database import async_session
    async with async_session() as db:
        devices = (await db.execute(select(Device).order_by(Device.id))).scalars().all()

    sem = asyncio.Semaphore(5)  # 并发上限（SNMP 引擎全局信号量 20 兜底）

    async def _sync_one(device_id: int, name: str, ip: str) -> dict:
        async with sem:
            async with async_session() as db:
                try:
                    changed, detail = await _sync_device_core(db, device_id)
                    return {
                        "device_id": device_id, "name": name, "ip": ip,
                        "success": True, "changed": changed, "detail": detail,
                    }
                except HTTPException as e:
                    try:
                        await db.rollback()
                    except Exception:
                        pass
                    return {
                        "device_id": device_id, "name": name, "ip": ip,
                        "success": False, "changed": [], "detail": e.detail,
                    }
                except Exception as e:
                    try:
                        await db.rollback()
                    except Exception:
                        pass
                    logger.warning(f"sync-all device {device_id} failed: {e}")
                    return {
                        "device_id": device_id, "name": name, "ip": ip,
                        "success": False, "changed": [], "detail": f"同步异常: {e}",
                    }

    results = await asyncio.gather(*[_sync_one(d.id, d.name, d.ip) for d in devices])
    ok = sum(1 for r in results if r["success"])
    # 汇总审计（独立会话）
    async with async_session() as db:
        await record_audit(
            db, {"sub": "system", "role": "system"}, "device", "sync_all",
            f"一键同步 {len(devices)} 台设备：成功 {ok} 台，失败 {len(results) - ok} 台",
        )
    return {
        "total": len(devices),
        "success": ok,
        "failed": len(results) - ok,
        "results": results,
    }


@router.post("/{device_id}/sync", response_model=DeviceResponse)
async def sync_device(device_id: int, db: AsyncSession = Depends(get_db)):
    """重新通过 SNMP 发现并同步设备信息（厂商/型号/类型/序列号/名称）。

    与创建时的「仅补全留空字段」不同，本接口会**用设备当前 SNMP 返回值覆盖**
    vendor / model / device_type / serial_number / name，解决「在设备上改名后平台
    不更新」的问题。设备不可达时返回 409 并保留原有信息。
    """
    await _sync_device_core(db, device_id)
    result = await db.execute(select(Device).where(Device.id == device_id))
    device = result.scalar_one_or_none()
    return DeviceResponse.model_validate(device)


async def _sync_device_core(db: AsyncSession, device_id: int) -> tuple[list[str], str]:
    """单台设备同步核心逻辑（厂商/型号/类型/序列号/名称覆盖 + 组件明细重建）。

    在传入的 db 会话内查询设备并执行同步，返回 (变更字段列表, 摘要)。
    设备不存在抛 404，设备不可达抛 409。
    """
    result = await db.execute(select(Device).where(Device.id == device_id))
    device = result.scalar_one_or_none()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    community = reveal_secret(device.snmp_community) or "aiops"
    try:
        info = await discover_device(device.ip, community)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"SNMP 发现失败: {e}")

    if not info.get("reachable"):
        raise HTTPException(status_code=409, detail=f"设备 {device.ip} SNMP 不可达，无法同步")

    # 采集实体 MIB 组件明细（板卡序列号/型号/版本等）
    components = []
    try:
        components = await collect_entity_components(device.ip, community)
    except Exception as e:
        logger.warning(f"collect entity components failed for {device.ip}: {e}")

    changed = []
    for field in ("vendor", "model", "device_type"):
        new_val = info.get(field)
        if new_val:
            old = getattr(device, field)
            if old != new_val:
                setattr(device, field, new_val)
                changed.append(field)
    # 序列号：优先机箱（phys_index=2），其次实体表第一个非空值
    chassis_serial = pick_chassis_serial(components) or info.get("serial_number")
    if chassis_serial and device.serial_number != chassis_serial:
        device.serial_number = chassis_serial
        changed.append("serial_number")
    # 名称：用设备当前 sysName 覆盖（即使之前是手动命名）
    new_name = info.get("name")
    if new_name and new_name != device.name:
        device.name = new_name
        changed.append("name")

    # 刷新组件明细（整表替换）
    from app.models.device_component import DeviceComponent
    from sqlalchemy import delete as sa_delete
    await db.execute(sa_delete(DeviceComponent).where(DeviceComponent.device_id == device.id))
    if components:
        db.add_all([
            DeviceComponent(
                device_id=device.id,
                phys_index=c.get("phys_index"),
                name=c.get("name"),
                descr=c.get("descr"),
                model_name=c.get("model_name"),
                serial_number=c.get("serial_number"),
                hardware_rev=c.get("hardware_rev"),
                firmware_rev=c.get("firmware_rev"),
                software_rev=c.get("software_rev"),
                mfg_name=c.get("mfg_name"),
            )
            for c in components
        ])

    await db.commit()
    await db.refresh(device)
    detail = f"同步设备 {device.name}({device.ip})"
    if changed:
        detail += f"：{'、'.join(changed)}"
    await record_audit(db, {"sub": "system", "role": "system"}, "device", "sync", detail)
    return changed, detail


@router.get("/{device_id}/components")
async def get_device_components(device_id: int, db: AsyncSession = Depends(get_db)):
    """返回设备硬件组件明细（实体 MIB 采集）。

    按「名称+序列号+型号」去重：部分设备（尤其模拟器）实体表会为每个
    槽位生成大量名称/序列号完全相同的容器行，去重后仅保留每类第一条，
    避免详情页出现几百行重复数据。
    """
    from app.models.device_component import DeviceComponent
    result = await db.execute(
        select(DeviceComponent)
        .where(DeviceComponent.device_id == device_id)
        .order_by(DeviceComponent.phys_index)
    )
    seen: set[tuple] = set()
    out = []
    for c in result.scalars().all():
        if not (c.serial_number or c.model_name or c.name):  # 过滤无信息的空容器
            continue
        if c.name and _PORT_ENTITY_RE.search(c.name):  # 排除物理接口/端口实体
            continue
        key = (c.name or "", c.serial_number or "", c.model_name or "")
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "phys_index": c.phys_index,
            "name": c.name,
            "descr": c.descr,
            "model_name": c.model_name,
            "serial_number": c.serial_number,
            "hardware_rev": c.hardware_rev,
            "firmware_rev": c.firmware_rev,
            "software_rev": c.software_rev,
            "mfg_name": c.mfg_name,
        })
    return out


# === 设备发现 ===

@router.post("/discover", response_model=DeviceDiscoverResponse)
async def discover_devices(data: DeviceDiscoverRequest, db: AsyncSession = Depends(get_db)):
    """通过 SNMP 发现设备信息，已纳管的设备标记为 already_managed"""
    # Get all already-managed IPs
    result = await db.execute(select(Device.ip))
    managed_ips = {row[0] for row in result.fetchall()}

    # Discover all IPs in parallel (limit concurrency to 10)
    semaphore = asyncio.Semaphore(10)
    async def limited_discover(ip):
        async with semaphore:
            return await discover_device(ip, data.snmp_community)

    tasks = [limited_discover(ip.strip()) for ip in data.ips if ip.strip()]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    discovered = []
    already_managed_count = 0
    for r in results:
        if isinstance(r, Exception) or not r.get("reachable"):
            continue
        is_managed = r["ip"] in managed_ips
        if is_managed:
            already_managed_count += 1
        discovered.append(DiscoveredDevice(
            ip=r["ip"],
            name=r.get("name"),
            vendor=r.get("vendor"),
            model=r.get("model"),
            device_type=r.get("device_type"),
            sys_descr=r.get("sys_descr"),
            already_managed=is_managed,
        ))

    return DeviceDiscoverResponse(
        total=len(discovered),
        discovered=discovered,
        already_managed_count=already_managed_count,
    )


@router.post("/batch", response_model=DeviceListResponse)
async def batch_create_devices(data: DeviceBatchCreate, db: AsyncSession = Depends(get_db)):
    """批量创建设备"""
    # Get existing IPs to skip duplicates
    result = await db.execute(select(Device.ip))
    existing_ips = {row[0] for row in result.fetchall()}

    # 计算本次实际新增数量（去重后），校验设备总数上限
    new_to_add = [d for d in data.devices if d.ip not in existing_ips]
    current_count = (await db.execute(select(func.count(Device.id)))).scalar() or 0
    if current_count + len(new_to_add) > settings.MAX_DEVICES:
        raise HTTPException(
            status_code=400,
            detail=(
                f"设备数量将超出上限（{settings.MAX_DEVICES} 台）。"
                f"当前已纳管 {current_count} 台，本次将新增 {len(new_to_add)} 台，"
                f"上限 {settings.MAX_DEVICES} 台。"
            ),
        )

    created = []
    for dev_data in new_to_add:
        device = Device(**protect_device_secrets(dev_data.model_dump()))
        db.add(device)
        created.append(device)

    await db.commit()
    for d in created:
        await db.refresh(d)

    # 自动通过 SNMP 补全各设备信息（仅填充留空字段）
    for d in created:
        await _enrich_device(db, d)

    return DeviceListResponse(total=len(created), items=[DeviceResponse.model_validate(d) for d in created])


async def _enrich_device(db: AsyncSession, device: Device):
    """创建设备后通过 SNMP 自动补全信息：仅填充用户留空的字段
    （vendor / model / device_type / serial_number / name）。
    设备 SNMP 不可达时静默跳过，不影响设备创建。"""
    community = reveal_secret(device.snmp_community) or "aiops"
    try:
        info = await discover_device(device.ip, community)
    except Exception as e:
        logger.warning(f"Enrich device {device.ip} failed: {e}")
        return
    if not info.get("reachable"):
        return
    # 仅补全用户未填写的字段
    for field in ("vendor", "model", "device_type", "serial_number"):
        cur = getattr(device, field)
        if (cur is None or cur == "") and info.get(field):
            setattr(device, field, info[field])
    # 设备名：用户未填或与 IP 相同（占位）时，用 SNMP sysName
    if (not device.name or device.name == device.ip) and info.get("name"):
        device.name = info["name"]
    try:
        await db.commit()
        await db.refresh(device)
    except Exception as e:
        logger.warning(f"Enrich commit failed for {device.ip}: {e}")
        await db.rollback()
